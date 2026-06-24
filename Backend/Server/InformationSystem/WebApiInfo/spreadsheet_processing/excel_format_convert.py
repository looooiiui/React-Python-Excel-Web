#====引入的 Openpyxl 库中的工作簿，工作表，以及加载函数====#
from openpyxl.worksheet.worksheet   import Worksheet        # 工作表类
from typing                         import Optional         # 多类型注释，用于函数返回多类型

from debug_tool.debug_util          import DebugTool        # 导入调试工具
from excel_processing               import excel_sheet_get_pure_max_col
from excel_processing               import excel_sheet_get_pure_max_row 

class ExcelFormatConversion:
    #====将格式转为形如 { "NAME" : {"PATH": "?"}(乱序) ...}====#
    #================可以选择表头名字, 前提存在=================#
    #====转换时主键名和附键名在同一列表，选取主键名从而建立映射表===#

    @staticmethod
    def convert_injection_dict(
        input_sheet: Optional[Worksheet], 
        head_name: str = "ID"
        ) -> dict:

        # ===================== 1. 输入校验 ===================== #

        if input_sheet is None:
            DebugTool.debug_log(f"Excel注入格式转换: 没有传入表格")
            return {}
        
        # ===================== 2. 获取有效数据范围 ===================== #
        # 纯数据最大行/列（遇到空行/空列自动停止）
        processed_max_row = excel_sheet_get_pure_max_row(input_sheet)
        processed_max_col = excel_sheet_get_pure_max_col(input_sheet)
        if processed_max_col == -1 or processed_max_row == -1:
            DebugTool.debug_log(f"Excel注入格式转换: 表格数据不符合规范")
            return {}

        # ===================== 3. 构建表头映射表 ===================== #
        # 映射格式为 {"model_name": index} 表头与其对应在Excel中的格式
        head_mapping: dict = {}
        for col_index in range(1, processed_max_col + 1):
            head_call_value = input_sheet.cell(row = 1, column = col_index).value

            # 空表头跳过
            if head_call_value is None:
                continue

            name_value: str = str(head_call_value).strip()

            # 重复表头停止映射，关闭程序
            if name_value in head_mapping:
                DebugTool.debug_log(f"Excel注入格式转换: 第{col_index}列表头'{name_value}'重复，停止映射")
                return {}

            head_mapping[name_value] = col_index
        
        # ===================== 4. 校验主键是否存在 ===================== #
        if head_name not in head_mapping:
            DebugTool.debug_log(f"Excel注入格式转换: 需求表头 {head_name} 不存在")
            return {}
        
        # ===================== 5. 逐行转换为字典 ===================== #
        convert_result: dict = {}
        # 填入转换字典(按照行转换),将对应的模块与其子属性对应
        # 从第2行开始遍历（第1行是表头）
        for row_index in range(2, processed_max_row + 1):
            attribute_dict: dict = {}

            # 获取表头列元素(主键)
            main_name_map_index: int = head_mapping[head_name]
            main_name = input_sheet.cell(row = row_index, column = main_name_map_index).value

            # 表头为空或者有 "\n"," " 等字符跳过
            if main_name is None or str(main_name).strip() == "":
                continue
            
            main_key = str(main_name).strip()

            # 将对应模块属性整理成字典(表头不一定在第一个,乱序)
            for attribute_name in head_mapping:
            
                # 主键本身不放入属性字典
                if attribute_name == head_name:
                    continue
                
                # 获取模块对应在Excel中的列索引位置，并取对应值
                attribute_index: int = head_mapping[attribute_name]
                attribute_value = input_sheet.cell(row = row_index, column = attribute_index).value

                # 空值统一转为 None
                if attribute_value is None or str(attribute_value).strip() == "":
                    attribute_value = None

                # 加入键值对(乱序)
                attribute_dict[attribute_name] = attribute_value
        
            # 存入最终结果
            convert_result[main_key] = attribute_dict

        DebugTool.debug_log(f"Excel注入格式转换: 注入字典转换成功")
        return convert_result
        
