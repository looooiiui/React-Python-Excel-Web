#====引入的 Openpyxl 库中的工作簿，工作表，以及加载函数====#
from debug_tool.debug_util          import DebugTool        # 导入调试工具
from openpyxl.worksheet.worksheet   import Worksheet        # 工作表类
from openpyxl.workbook.workbook     import Workbook         # 工作簿类
from openpyxl                       import load_workbook    # 加载Excel函数
from typing                         import Optional         # 多类型注释，用于函数返回多类型



"""
全局Excel文件管理器
可以替换所有函数内打开Excel的操作
减少频繁打开关闭Excel的性能消耗
这里可以忽略，本文并未使用
"""
#===============全局文件管理器Excel===============#
"""
统一管理全局被打开的Excel
防止手动释放出现问题
"""
class ExcelManager:

    #==默认空表未打开Excel==#
    def __init__(self, file_path: Optional[str] = None):

        self._inner_excel_path  : Optional[str]         = file_path       # 记录Excel路径
        self._inner_excel       : Optional[Workbook]    = None            # 记录Excel文件
        #==尝试打开Excel路径==#
        if self._inner_excel_path is not None:
            try:
                read_excel: Workbook = load_workbook(
                                                    file_path, 
                                                    read_only=True
                                                    )
                #==打开则记录 Excel 到当前类对象==#
                self._inner_excel = read_excel
                DebugTool.debug_log(f"Excel统一类: 初始化Excel成功")

            #=====异常处理=====#
            except Exception as e:
                DebugTool.debug_log(f"Excel统一类: 初始化失败: {e}")
        else:
            DebugTool.debug_log(f"创建空Excel统一类成功")

    #==尝试打开Excel==#
    def open_excel(self, file_path: Optional[str] = None) -> bool: 
        if file_path is None:
            DebugTool.debug_log(f"Excel统一类: 没有传入打开路径")
            return False
            
        #==检查本对象是否已经有Excel指向==#
        if self._inner_excel:
            DebugTool.debug_log(f"Excel统一类: 打开对象已存在: {self._inner_excel}: 请手动释放")
            return False

        #==尝试读取Excel==#
        try:
            #==获取Excel,同时更新内部状态==#
            read_excel: Workbook        = load_workbook(file_path, read_only=True)
            self._inner_excel           = read_excel
            self._inner_excel_path      = file_path
            return True
    
        #==异常处理==#
        except Exception as e:
            DebugTool.debug_log(f"Excel统一类: 手动打开Excel发生异常: 打开失败")
            return False

    #==获得类中Excel==#
    def get_excel(self) -> Optional[Workbook]:
        #==检查本对象Excel是否已经指向==#
        if self._inner_excel:
            DebugTool.debug_log(f"Excel统一类: 成功返回Excel")
            return self._inner_excel
        
        #=不存在=#
        DebugTool.debug_log(f"Excel统一类: 返回excel失败,excel并未读取")
        return None

    #==获得Excel中对应表单==#   
    def get_sheet(self, sheet_name: Optional[str] = None) -> Optional[Worksheet]:
        if sheet_name is None:
            DebugTool.debug_log(f"Excel统一类: 没有传入表单名称")
            return None         

        #==检查Excel指向==#
        if self._inner_excel is None:
            DebugTool.debug_log(f"Excel统一类: 不存在已经打开的Excel表")
            return None

        #==返回特定表==#     
        if sheet_name not in self._inner_excel.sheetnames:
            DebugTool.debug_log(f"Excel统一类: 工作表'{sheet_name}' 不存在")
            return None
        #==返回表==#
        DebugTool.debug_log(f"Excel统一类: 工作表返回成功")
        return self._inner_excel[sheet_name]

    #==关闭Excel工作簿以及停止指向==#
    def close_workbook(self) -> bool:
        #==检查Excel指向==#
        if self._inner_excel is None:
            DebugTool.debug_log(f"Excel统一类: 不存在原表,已经关闭")
            return True
        
        # 尝试关闭表格
        try:
            self._inner_excel.close()
            DebugTool.debug_log(f"Excel统一类: 工作簿关闭成功")
            return True
        
        #==表单关闭失败==#
        except Exception as e:
            DebugTool.debug_log(f"Excel统一类: 关闭excel出现异常: {e}")
            return False   
        
        #======无论结果置空内部表格存储指向========#
        finally:
            self._inner_excel       = None
            self._inner_excel_path  = None 


#===============以下所有工具使用失败均返回 -1 =============#

#==============获取去除杂乱数据后的表格的最大行数===========#
#=============即找到第一个为空的格子(行或者列)记为最大值=====#
#=======================包含开头=========================#
#===========参数: (表格，基于第几列为遍历基础) =============#
def excel_sheet_get_pure_max_row(input_sheet: Optional[Worksheet], col_index: int = 1) -> int:
    if input_sheet is None:
        DebugTool.debug_log(f"纯净最大行: 没有传入表格")
        return -1

    #=======寻找到名字中断则返回纯净最大值======#
    try:
        max_row_length: int         = input_sheet.max_row
        # 检查表格是否有表头对应表格数据
        if max_row_length < 2:
            DebugTool.debug_log(f"纯净最大行: 表格数据为空或只有标题")
            return -1
        
        # 验证第一个单元格是否为空，为空则认为表格不合规范
        cell_value = input_sheet.cell(row = 1, column = col_index).value
        if cell_value is None or str(cell_value).strip() == "":
            DebugTool.debug_log(f"表格第 {col_index} 列第一行为空")
            return -1
        
        # 遍历寻找最大值
        for index in range(2, max_row_length + 1):
            cell_value = input_sheet.cell(row = index, column = col_index).value
            # print(cell_value)
            if cell_value is None or str(cell_value).strip() == "":
                return index - 1
            
        # 全连续返回最大值
        return max_row_length
    except Exception as e:
        DebugTool.debug_log(f"纯净最大行: 出现未知问题: {e}")
        return -1

#===========获取去除杂乱数据后的表格的最大列数==========#
def excel_sheet_get_pure_max_col(input_sheet: Optional[Worksheet], row_index: int = 1) -> int:
    if input_sheet is None:
        DebugTool.debug_log(f"纯净最大列: 没有传入表格")
        return -1
    
    # 验证第一个单元格是否为空，为空则认为表格不合规范
    cell_value = input_sheet.cell(row = row_index, column = 1).value
    if cell_value is None or str(cell_value).strip() == "":
        DebugTool.debug_log(f"表格第 {row_index} 行第一列为空")
        return -1

    #=======寻找到名字中断则返回纯净最大值======#
    try:
        max_col_length: int         = input_sheet.max_column
        if max_col_length < 1:
            DebugTool.debug_log(f"纯净最大列: 表格数据没有列")
            return -1
        # 遍历寻找最大值
        for index in range(1, max_col_length + 1):
            cell_value = input_sheet.cell(row = row_index, column = index).value
            # print(cell_value)
            if cell_value is None or str(cell_value).strip() == "":
                return index - 1
            
        # 全连续返回最大值
        return max_col_length
    except Exception as e:
        DebugTool.debug_log(f"纯净最大列: 出现未知问题: {e}")
        return -1     








